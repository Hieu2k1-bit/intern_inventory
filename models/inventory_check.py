# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.
import xlsxwriter

from odoo import fields, models, api, _
import base64
import io
from io import BytesIO
from openpyxl.reader.excel import load_workbook

try:
    import openpyxl
except ImportError:
    openpyxl = None


class InventoryCheck(models.Model):
    _name = 'inventory.check'
    _description = 'Inventory Check Sheet'

    name = fields.Text(string='Check Name')
    warehouse_id = fields.Many2one('stock.warehouse', string='Warehouse ID')
    location_id = fields.Many2one('stock.location', string='Warehouse Location ID')
    employeeCheck_id = fields.Many2one('res.users', string='Employee Check')
    company = fields.Many2one('res.company', string='Company')
    state = fields.Selection([('ready', 'Ready'), ('done', 'Done')], default='ready', string='State')
    create_datetime = fields.Datetime(string='Create Date', required=True, default=lambda self: fields.Datetime.now())
    check_date = fields.Date(string='Check Date', required=True, default=lambda self: fields.Date.today())
    line_ids = fields.One2many('inventory.line', 'check_id', string='Product Check Lines')
    display_warehouse_name = fields.Char(string='Warehouse', compute='_compute_display_warehouse_name', store=True)
    display_warehouse_location = fields.Char(string='Warehouse Location', compute='_compute_display_warehouse_location',
                                             store=True)

    @api.depends('warehouse_id')
    def _compute_display_warehouse_name(self):
        for rec in self:
            if rec.warehouse_id:
                rec.display_warehouse_name = rec.warehouse_id.name
            else:
                rec.display_warehouse_name = "All Warehouse"

    @api.depends('location_id')
    def _compute_display_warehouse_location(self):
        for rec in self:
            if rec.location_id:
                rec.display_warehouse_location = rec.location_id.name
            else:
                rec.display_warehouse_location = "All Location"

    # def _compute_display_employeeCheck(self):
    #     for rec in self:
    #         if rec.employeeCheck_id:
    #             rec.display_arehouswe_location = rec.location_id.name
    #         else:
    #             rec.display_warehouse_location = "All Location"

    def action_complete(self):
        self.ensure_one()  # Dam bao chi xu ly mot form mot luc
        self.write({'state': 'done',
                    'create_datetime': fields.Datetime.now()})  # Tu dong luu cac thay doi
        return True


    def action_apply_all(self):
        self.ensure_one()
        for line in self.line_ids:
            line.diff_quantity = abs((line.quantity or 0.0) - (line.inventory_quantity or 0.0))


    def action_open_product_selection(self):
        self.ensure_one()
        return {
            'name': 'Chose Products',
            'type': 'ir.actions.act_window',
            'res_model': 'product.selection.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'default_check_id': self.id,
            }
        }

    def import_data(self, file):
        if not openpyxl:
            return {'status': 'error', 'message': _('Library openpyxl is not installed.')}
        try:
            file_content = base64.b64decode(file)
            data = BytesIO(file_content)
            workbook = load_workbook(filename=data)
            if 'Inventory sheet' not in workbook.sheetnames:
                return {
                    'status': 'error',
                    'message': _("Sheet 'Inventory sheet' not found in Excel file")
                }
            project_sheet = workbook['Inventory sheet']
            # Extract headers and rows for 'Inventory sheet' sheet
            project_header = [
                cell.value.strip() if isinstance(cell.value, str) else cell.value
                for cell in next(project_sheet.iter_rows(min_row=1, max_row=1))
            ]  # get the first row value of the cel (header)
            project_rows = list(project_sheet.iter_rows(min_row=2, values_only=True))  #get value in cell instead of cel
            print('project_header', project_header)
            print('Number of project rows', project_rows)

            # Required fields for 'Inventory sheet' sheet
            required_project_field = ['Inventory sheet', 'employeeCheck', 'Production']
            for field in required_project_field:
                if field not in project_header:
                    return {
                        'status': 'error',
                        'message': _(f"Thiếu cột bắt buộc: '{field}'Missing required column: in sheet'Inventory sheet'")
                    }
            # Header mapping for 'Inventory sheet' sheet
            project_header_mapping = {
                'Inventory sheet': 'name',
                'Inventory date': 'check_date',
                'employeeCheck': 'employeeCheck_id',
                'warehouse': 'warehouse_id',
                'location': 'location_id',
                'product': 'product.id',
                'Lot/serial number': 'lot_id',
                'Unit': 'unit_id',
                'quantity': 'quantity',
                'Quantity counted': 'inventory_quantity',
            }
            # Collect all project codes from 'Phiếu kiểm kê' sheet
            project_codes = set()
            for row in project_rows:
                row_data = {project_header[i]: row[i] for i in range(len(project_header)) if i < len(row)}
                code = row_data.get('Inventory sheet')
                if code is not None:
                    project_codes.add(str(code).strip())  # convert to string and strip
                else:
                    return {
                        'status': 'error',
                        'message': f"Thiếu 'Phiếu kiểm kê' trong dòng dữ liệu của sheet 'Phiếu kiểm kê': {row_data}"
                    }
            projects = {}
            for row in project_rows:
                row_data = {project_header[i]: row[i] for i in range(len(project_header)) if i < len(row)}
                project_data = self._prepare_project_data(row_data, project_header_mapping)
                if isinstance(project_data, dict) and project_data.get('status') == 'error':
                    return project_data
                project = self._update_or_create_project(project_data)
                if isinstance(project, dict) and project.get('status') == 'error':
                    return project
                projects[project_data['name']] = project
            return {
                'status': 'success',
                'message': 'Nhập dữ liệu thành công!'
            }
        except Exception as e:
            self.env.cr.rollback()
            return {
                "status": "error",
                "message": f"Lỗi khi nhập dữ liệu: {str(e)}"
            }

    def action_export_excel(self):
        for rec in self:
            output = BytesIO()
            workbook = xlsxwriter.Workbook(output, {'in_memory': True})
            sheet = workbook.add_worksheet('Kiểm kê kho')

            sheet.set_column(0, 0, 20)
            sheet.set_column(1, 1, 15)
            sheet.set_column(2, 2, 20)
            sheet.set_column(3, 3, 15)
            sheet.set_column(4, 4, 20)
            sheet.set_column(5, 5, 25)
            sheet.set_column(6, 6, 20)
            sheet.set_column(7, 7, 10)
            sheet.set_column(8, 8, 15)
            sheet.set_column(9, 9, 15)

            header_format = workbook.add_format({
                'bold': True,
                'font_color': 'red',
                'align': 'center',
                'valign': 'vcenter',
                'bg_color': '#F9F9F9',
                'border': 1
            })

            headers = [
                'Phiếu kiểm kê', 'Ngày kiểm kê', 'Người kiểm kê',
                'Kho', 'Vị trí', 'Sản phẩm', 'Số lô/serial',
                'ĐVT', 'Số lượng hiện có', 'Số lượng đã đếm'
            ]
            for col_num, header in enumerate(headers):
                sheet.write(0, col_num, header, header_format)

            row = 1
            for line in rec.line_ids:
                sheet.write(row, 0, rec.name or '')
                sheet.write(row, 1, str(rec.inventory_date or ''))
                sheet.write(row, 2, rec.user_id.name or '')
                sheet.write(row, 3, rec.warehouse_id.name if rec.warehouse_id else 'Tất cả kho')
                sheet.write(row, 4, line.location_id.display_name or line.location_id.complete_name or '')
                sheet.write(row, 5, line.product_id.name or '')
                sheet.write(row, 6, line.lot_id.name if line.lot_id else '')
                sheet.write(row, 7, line.uom_id.name if line.uom_id else '')
                sheet.write(row, 8, line.quantity or 0)
                sheet.write(row, 9, line.quantity_counted or 0)
                row += 1

            workbook.close()
            output.seek(0)
            data = output.read()

            export_attachment = self.env['ir.attachment'].create({
                'name': f'PhieuKiemKe_{rec.name}.xlsx',
                'type': 'binary',
                'datas': base64.b64encode(data),
                'res_model': 'intern_inventory.check',
                'res_id': rec.id,
            })

            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/{export_attachment.id}?download=true',
                'target': 'self',
            }

    def _prepare_project_data(self, row_data, mapping):
        result = {}
        for excel_col, field_name in mapping.items():
            value = row_data.get(excel_col)
            if not value:
                continue
            if field_name in ['user_id', 'warehouse_id', 'location_id', 'uom_id', 'lot_id', 'product.id']:
                model = {
                    'employeeCheck_id': 'res.users',
                    'warehouse_id': 'stock.warehouse',
                    'location_id': 'stock.location',
                    'unit_id': 'uom.uom',
                    'lot_id': 'stock.lot',
                    'product.id': 'product.product',
                }[field_name]
                record = self.env[model].search([('name', '=', value)], limit=1)
                if not record:
                    return {
                        'status': 'error',
                        'message': f"Không tìm thấy '{value}' trong '{excel_col}' (model {model})"
                    }
                result[field_name if field_name != 'product.id' else 'product_id'] = record.id
            else:
                result[field_name] = value
        result['name'] = str(row_data.get('Phiếu kiểm kê')).strip()  # dùng làm mã kiểm kê duy nhất
        return result

    def _update_or_create_project(self, vals):
        # Kiểm tra nếu đã có thì cập nhật, nếu chưa thì tạo mới
        existing = self.search([('name', '=', vals.get('name'))], limit=1)
        if existing:
            existing.write(vals)
            return existing
        else:
            return self.create(vals)
