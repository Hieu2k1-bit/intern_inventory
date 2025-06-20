import base64
from io import BytesIO
from odoo import models, fields, api, _
from openpyxl.reader.excel import load_workbook
from reportlab.graphics.shapes import translate
import xlsxwriter
from odoo.exceptions import UserError

try:
    import openpyxl
except ImportError:
    openpyxl = None


class InternInventory(models.Model):
    _name = 'intern_inventory.check'
    _description = 'Inventory Check Sheet'

    name = fields.Char(string='Name', required=True)

    warehouse_ids = fields.Many2many('stock.warehouse', string="Warehouse")
    location_ids = fields.Many2many('stock.location', string="Location")
    user_id = fields.Many2one('res.users', string="Responsible Person", default=lambda self: self.env.user)
    create_date = fields.Datetime(string="Create date", required=True, default=fields.Datetime.now)
    inventory_date = fields.Date(string="Inventory date", required=True, default=fields.Date.today)
    state = fields.Selection([
        ('ready', 'READY'),
        ('done', 'DONE'),
    ], string='Status', default='ready', readonly=True)
    company_name = fields.Char(string="Company", default="My Company1")

    warehouse_display_name = fields.Char(
        string='Warehouse',
        compute='_compute_warehouse_display_name',
        store=True
    )

    location_display_name = fields.Char(
        string='Location',
        compute='_compute_location_display_name',
        store=True
    )
    state_display_name = fields.Char(
        string='Status',
        compute='_compute_state_display_name',
        store=True
    )
    product_id = fields.Many2one('product.product', string="Product List")
    lot_id = fields.Many2one('stock.lot', string="Lot/Serial Number")
    uom_id = fields.Many2one('uom.uom', string="Unit of Measure")
    quantity = fields.Float(string="Available Quantity")
    quantity_counted = fields.Float(string="Quantity Counted")
    line_ids = fields.One2many('inventory.line', 'check_id', string='Product Check Lines')

    @api.depends('state')
    def _compute_state_display_name(self):
        for rec in self:
            if rec.state == 'ready':
                rec.state_display_name = "Ready"
            else:
                rec.state_display_name = "Done"

    @api.depends('warehouse_ids')
    def _compute_warehouse_display_name(self):
        for record in self:
            names = record.warehouse_ids.mapped('name')
            record.warehouse_display_name = ', '.join(names) if names else _("All Warehouses")

    @api.depends('location_ids')
    def _compute_location_display_name(self):
        for record in self:
            names = record.location_ids.mapped('name')
            record.location_display_name = ', '.join(names) if names else _("All Locations")

    @api.onchange('warehouse_ids')
    def _onchange_warehouse_ids(self):
        if self.warehouse_ids:
            warehouse_view_locations = self.warehouse_ids.mapped('view_location_id')
        else:
            # Lấy tất cả các view location của tất cả kho
            all_warehouses = self.env['stock.warehouse'].search([])
            warehouse_view_locations = all_warehouses.mapped('view_location_id')

        all_locations = self.env['stock.location'].search([
            ('location_id', 'child_of', warehouse_view_locations.ids),
            ('usage', '=', 'internal')
        ])
        return {
            'domain': {'location_ids': [('id', 'in', all_locations.ids)]}
        }

    def action_ready(self):
        self.ensure_one()
        self.write({'state': 'ready'})

    def action_done(self):
        self.ensure_one()
        self.write({'state': 'done'})

    def get_warehouse_from_location(self, loc):
        current_loc = loc
        while current_loc and current_loc.usage != 'view':
            current_loc = current_loc.location_id
        if current_loc:
            return self.env['stock.warehouse'].search([('view_location_id', '=', current_loc.id)], limit=1)
        return False

    def import_data(self, file):
        self.ensure_one()
        if self.location_ids and not self.warehouse_ids:
            return {
                'status': 'error',
                'message': _("You must select the warehouse once you have selected the location!")
            }
        if self.state != 'ready':
            raise UserError(_("Import is only allowed when the status is 'Ready'."))
        if not openpyxl:
            return {'status': 'error', 'message': _('Library openpyxl is not installed.')}
        try:
            file_content = base64.b64decode(file)
            data = BytesIO(file_content)
            workbook = load_workbook(filename=data)
            if 'Phiếu kiểm kê' not in workbook.sheetnames:
                return {
                    'status': 'error',
                    'message': _("Sheet 'Phiếu kiểm kê' not found in Excel file")
                }
            project_sheet = workbook['Phiếu kiểm kê']
            # Extract headers and rows for 'Phiếu kiểm kê' sheet
            # get the first row value of the cel (header)
            project_header = [
                str(cell.value or '').strip()
                for cell in next(project_sheet.iter_rows(min_row=1, max_row=1))
            ]
            project_rows = list(project_sheet.iter_rows(min_row=2, values_only=True))  # lấy giá trị trong ô thay vì cel
            print('project_header', project_header)
            print('Number of project rows', project_rows)

            # Required fields for 'Phiếu kiểm kê' sheet
            required_project_field = ['Kho', 'Vị trí', 'Sản phẩm']
            for field in required_project_field:
                if field not in project_header:
                    return {
                        'status': 'error',
                        'message': _(f"Thiếu cột bắt buộc: '{field}' trong sheet'Phiếu kiểm kê'")
                    }
            # Header mapping for 'Phiếu kiểm kê' sheet
            project_header_mapping = {
                'Kho': 'warehouse_ids',
                'Vị trí': 'location_ids',
                'Sản phẩm': 'product_id',
                'Số lô/serial': 'lot_id',
                'ĐVT': 'uom_id',
                'Số lượng hiện có': 'quantity',
                'Số lượng đã đếm': 'quantity_counted',
            }
            # Group data by 'Inventory sheet'
            line_vals = []
            processed_products = set()
            processed_lines = set()
            valid_line_count = 0
            for row in project_rows:
                # Create row_data dict from Excel row
                row_data = {}
                for i in range(len(project_header)):
                    if i < len(row):
                        header_key = project_header[i]
                        cell_value = row[i]
                        row_data[header_key] = str(cell_value or '').strip()
                    else:
                        row_data[project_header[i]] = ''

                # --- Warehouse ---
                warehouse = False
                warehouse_name_from_excel = row_data.get('Kho')
                warehouse_name_from_excel = warehouse_name_from_excel.strip() if warehouse_name_from_excel else ''

                if self.warehouse_ids:
                    # If the user selects a specific warehouse, the Excel row must have warehouse and belong to warehouse_ids
                    if not warehouse_name_from_excel or warehouse_name_from_excel not in self.warehouse_ids.mapped(
                            'name'):
                        continue

                if warehouse_name_from_excel:
                    warehouse = self.env['stock.warehouse'].search([('name', '=', warehouse_name_from_excel)], limit=1)
                    if not warehouse:
                        return {'status': 'error', 'message': f"No warehouse found: '{warehouse_name_from_excel}'."}

                # --- Location ---
                location = False
                location_name_from_excel = row_data.get('Vị trí')
                location_name_from_excel = location_name_from_excel.strip() if location_name_from_excel else ''

                if self.location_ids:
                    # If the user selects a specific location, the Excel row must have a location and belong to location_ids
                    if not location_name_from_excel or location_name_from_excel not in self.location_ids.mapped('name'):
                        continue
                else:
                    # If the user does not select any location, all locations are considered selected.
                    pass  # no further testing required

                location_provided_in_excel = False
                if location_name_from_excel:
                    location_provided_in_excel = True
                    location_domain = [('name', '=', location_name_from_excel)]

                    if warehouse:
                        warehouse_view_location = warehouse.view_location_id
                        if warehouse_view_location:
                            location_domain.append(('location_id', 'child_of', warehouse_view_location.id))
                        else:
                            return {'status': 'error', 'message': f"Warehouse '{warehouse.name}' no main position."}

                    location = self.env['stock.location'].search(location_domain, limit=1)
                    if not location:
                        warehouse_part = f"In stock '{warehouse.name}' if warehouse else ''"
                        return {'status': 'error',
                                'message': f"Location not found: '{location_name_from_excel}' {warehouse_part}."}

                # Kiểm tra vị trí thuộc kho không
                if warehouse and location_provided_in_excel and location:
                    warehouse_of_location = self.get_warehouse_from_location(location)
                    if not warehouse_of_location or warehouse_of_location.id != warehouse.id:
                        return {'status': 'error',
                                'message': f"Location '{location.name}' is not in the warehouse '{warehouse.name}'."}

                # --- Xử lý sản phẩm ---
                product_name = row_data.get('Sản phẩm')
                if not product_name or not isinstance(product_name, str):
                    return {'status': 'error', 'message': _("Invalid or missing product name.")}
                product_key = product_name.strip()
                processed_products.add(product_key)
                product = self.env['product.product'].search([('default_code', '=', product_name)], limit=1)
                if not product:
                    return {'status': 'error', 'message': f"Product code not found: {product_name}"}

                if not warehouse:
                    return {
                        'status': 'error',
                        'message': f"Product '{product.name}' not in stock '{warehouse_name_from_excel}' or invalid warehouse name."
                    }
                if not self.warehouse_ids and warehouse:
                    internal_locations = self.env['stock.location'].search([
                        ('usage', '=', 'internal'),
                        ('location_id', 'child_of', warehouse.view_location_id.id)
                    ])
                    product_quants = self.env['stock.quant'].search([
                        ('product_id', '=', product.id),
                        ('location_id', 'in', internal_locations.ids),
                        ('quantity', '>', 0)
                    ], limit=1)

                    if not product_quants:
                        return {
                            'status': 'error',
                            'message': f"Product '{product.name}' does not exist in warehouse '{warehouse.name}'."
                        }


                combo_key = (product_key, warehouse.id if warehouse else False, location.id if location else False)
                if combo_key in processed_lines:
                    continue
                processed_lines.add(combo_key)

                # Tìm lot, uom
                lot_name = row_data.get('Số lô/serial')
                lot = False
                if lot_name and str(lot_name).strip():
                    lot = self.env['stock.lot'].search([('name', '=', str(lot_name).strip())], limit=1)
                uom_name = row_data.get('ĐVT')
                uom = False
                if uom_name and str(uom_name).strip():
                    uom = self.env['uom.uom'].search([('name', '=', str(uom_name).strip())], limit=1)

                # Nếu ô ĐVT trống hoặc không tìm thấy, lấy đơn vị tính mặc định của sản phẩm
                if not uom and product.uom_id:
                    uom = product.uom_id

                # Tìm quant
                quant_domain = [('product_id', '=', product.id)]
                if lot:
                    quant_domain.append(('lot_id', '=', lot.id))

                if location:
                    quant_domain.append(('location_id', '=', location.id))
                elif self.warehouse_ids:
                    all_locations = self.env['stock.location'].search([
                        ('warehouse_id', 'in', self.warehouse_ids.ids),
                        ('usage', '=', 'internal')
                    ])
                    if all_locations:
                        quant_domain.append(('location_id', 'in', all_locations.ids))
                    else:
                        return {'status': 'error',
                                'message': f"No internal locations found in the warehouse: {warehouse.name}."}

                quant = self.env['stock.quant'].search(quant_domain, limit=1)
                if not quant:
                    error_message = f"Sản phẩm '{product.name}'"
                    if lot:
                        error_message += f" (Số lô/serial: '{lot.name}')"
                    if location:
                        error_message += f" not in location '{location.name}'."
                    elif warehouse:
                        error_message += f" not in warehouse '{warehouse.name}'."
                    else:
                        error_message += f" does not exist in any location."
                    return {'status': 'error', 'message': error_message}

                # Thêm dòng vào line_vals
                line_vals.append((0, 0, {
                    'product_id': product.id,
                    'lot_id': lot.id if lot else False,
                    'uom_id': uom.id if uom else False,
                    'quant_id': quant.id if quant else False,
                    'quantity': quant.quantity if quant else 0,
                    'quantity_counted': row_data.get('Số lượng đã đếm') or 0,
                    'location_id': location.id if location else False,
                }))
                valid_line_count += 1
            if valid_line_count == 0:
                return {'status': 'error',
                        'message': "There are no rows in the Excel file that match the Warehouse and Location selected in the inventory sheet."}

            # Chuẩn bị dữ liệu phiếu kiểm kê từ dòng đầu
            first_row_data = {}
            for i in range(len(project_header)):
                if i < len(project_rows[0]):
                    header_key = project_header[i]
                    cell_value = project_rows[0][i]
                    first_row_data[header_key] = str(cell_value or '').strip()
                else:
                    first_row_data[project_header[i]] = ''

            main_data = self._prepare_project_data(first_row_data, project_header_mapping)
            if isinstance(main_data, dict) and main_data.get('status') == 'error':
                return main_data
            if not main_data.get('name'):
                main_data['name'] = self.name
            # Xóa kho và vị trí khỏi dữ liệu để không tự động ghi
            main_data.pop('warehouse_ids', None)
            main_data.pop('location_ids', None)
            # Viết dữ liệu phiếu kiểm kê
            self.write(main_data)
            self.write({'line_ids': line_vals})

            return {
                'status': 'success',
                'message': 'Data import successful!'
            }
        except Exception as e:
            self.env.cr.rollback()
            return {
                "status": "error",
                "message": f"Error while entering data: {str(e)}"
            }

    def _prepare_project_data(self, row_data, mapping):
        result = {}
        for excel_col, field_name in mapping.items():
            value = row_data.get(excel_col)
            if not value:
                continue
            if field_name in ['warehouse_ids', 'location_ids', 'uom_id', 'lot_id', 'product_id']:
                model = {
                    'warehouse_ids': 'stock.warehouse',
                    'location_ids': 'stock.location',
                    'uom_id': 'uom.uom',
                    'lot_id': 'stock.lot',
                    'product_id': 'product.product',
                }[field_name]
                model_env = self.env[model]
                if field_name == 'product_id':
                    domain = [('default_code', '=', str(value))]
                else:
                    domain = [('name', '=', str(value))]
                record = model_env.search(domain, limit=1)
                if not record:
                    return {
                        'status': 'error',
                        'message': f"Not found '{value}' in '{excel_col}' (model {model})"
                    }
                result[field_name] = record.id
            else:
                result[field_name] = value
        inventory_check_name = row_data.get('Phiếu kiểm kê')
        result['name'] = str(
            inventory_check_name).strip() if inventory_check_name is not None else ''  # dùng làm mã kiểm kê duy nhất
        return result

    # Bảo
    # def action_open_product_selection(self):
    #     self.ensure_one()
    #     self.env['product.selection.line.wizard'].search([('check_id', '=', self.id)]).unlink()
    #
    #     # Khởi tạo domain quants
    #     quant_domain = [('quantity', '>', 0)]
    #
    #     if self.location_id:
    #         quant_domain.append(('location_id', '=', self.location_id.id))
    #     elif self.warehouse_id:
    #         # Lấy tất cả các location thuộc warehouse
    #         warehouse_locations = self.env['stock.location'].search([
    #             ('id', 'child_of', self.warehouse_id.view_location_id.id)
    #         ])
    #         quant_domain.append(('location_id', 'in', warehouse_locations.ids))
    #     # else: không cần thêm điều kiện gì thêm, sẽ lấy tất cả tồn kho
    #
    #     # Truy vấn các quants phù hợp
    #     quants = self.env['stock.quant'].search(quant_domain)
    #
    #     for quant in quants:
    #         self.env['product.selection.line.wizard'].create({
    #             'product_id': quant.product_id.id,
    #             'internal_reference': quant.product_id.default_code,
    #             'location_id': quant.location_id.id,
    #             'check_id': self.id,
    #         })
    #
    #     return {
    #         'name': 'Chọn sản phẩm từ tồn kho',
    #         'type': 'ir.actions.act_window',
    #         'res_model': 'product.selection.line.wizard',
    #         'view_mode': 'tree,form',
    #         'domain': [('check_id', '=', self.id)],
    #         'target': 'new',
    #         'context': {
    #             'default_check_id': self.id,
    #         },
    #     }
    #
    # def action_export_excel(self):
    #     for rec in self:
    #         output = BytesIO()
    #         workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    #         sheet = workbook.add_worksheet('Kiểm kê kho')
    #
    #         sheet.set_column(0, 0, 20)
    #         sheet.set_column(1, 1, 15)
    #         sheet.set_column(2, 2, 20)
    #         sheet.set_column(3, 3, 15)
    #         sheet.set_column(4, 4, 20)
    #         sheet.set_column(5, 5, 25)
    #         sheet.set_column(6, 6, 20)
    #         sheet.set_column(7, 7, 10)
    #         sheet.set_column(8, 8, 15)
    #         sheet.set_column(9, 9, 15)
    #
    #         header_format = workbook.add_format({
    #             'bold': True,
    #             'font_color': 'red',
    #             'align': 'center',
    #             'valign': 'vcenter',
    #             'bg_color': '#F9F9F9',
    #             'border': 1
    #         })
    #
    #         headers = [
    #             'Phiếu kiểm kê', 'Ngày kiểm kê', 'Người kiểm kê',
    #             'Kho', 'Vị trí', 'Sản phẩm', 'Số lô/serial',
    #             'ĐVT', 'Số lượng hiện có', 'Số lượng đã đếm'
    #         ]
    #         for col_num, header in enumerate(headers):
    #             sheet.write(0, col_num, header, header_format)
    #
    #         row = 1
    #         for line in rec.line_ids:
    #             sheet.write(row, 0, rec.name or '')
    #             sheet.write(row, 1, str(rec.inventory_date or ''))
    #             sheet.write(row, 2, rec.user_id.name or '')
    #             sheet.write(row, 3, rec.warehouse_id.name if rec.warehouse_id else 'Tất cả kho')
    #             sheet.write(row, 4, line.location_id.display_name or line.location_id.complete_name or '')
    #             sheet.write(row, 5, line.product_id.name or '')
    #             sheet.write(row, 6, line.lot_id.name if line.lot_id else '')
    #             sheet.write(row, 7, line.uom_id.name if line.uom_id else '')
    #             sheet.write(row, 8, line.quantity or 0)
    #             sheet.write(row, 9, line.quantity_counted or 0)
    #             row += 1
    #
    #         workbook.close()
    #         output.seek(0)
    #         data = output.read()
    #
    #         export_attachment = self.env['ir.attachment'].create({
    #             'name': f'PhieuKiemKe_{rec.name}.xlsx',
    #             'type': 'binary',
    #             'datas': base64.b64encode(data),
    #             'res_model': 'intern_inventory.check',
    #             'res_id': rec.id,
    #         })
    #
    #         return {
    #             'type': 'ir.actions.act_url',
    #             'url': f'/web/content/{export_attachment.id}?download=true',
    #             'target': 'self',
    #         }
    #
    # # Đạt
    # def action_complete(self):
    #     self.ensure_one()  # Dam bao chi xu ly mot form mot luc
    #     self.write({'state': 'done',
    #                 'create_date': fields.Datetime.now()})  # Tu dong luu cac thay doi
    #     return True
    #     # self.env.ref('intern_inventory.action_inventory_check').read())[0] # Save and move to list view
    #
    # def action_apply_all(self):
    #     self.ensure_one()
    #     for line in self.line_ids:
    #         line.diff_quantity = abs((line.quantity or 0.0) - (line.quantity_counted or 0.0))
    #
